from pathlib import Path

from openpyxl import load_workbook

POST_GO_LIVE_SHEET = "Post-Go-Live Monitoring Report"


def load_post_go_live_rows() -> list[dict]:
    workbook_path = Path(__file__).with_name("monitoring_report.xlsx")
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[POST_GO_LIVE_SHEET]

    headers = [cell.value for cell in worksheet[1]]
    rows: list[dict] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[1] is None:
            continue
        item = dict(zip(headers, row))
        rows.append(item)

    return rows
